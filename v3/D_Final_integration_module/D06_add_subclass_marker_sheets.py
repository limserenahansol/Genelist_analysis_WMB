#!/usr/bin/env python
"""
D06_add_subclass_marker_sheets.py

Fixes the marker granularity bug in FINAL_decision.xlsx.

BEFORE: cell_type_marker_genes came from curated_marker_template.csv, which is
        keyed on cell_type_label. Every Allen subclass anchor under one cell
        type therefore showed an IDENTICAL marker list, so nothing separated
        e.g. 012 MEA Slc17a7 Glut from 113 MEA-COA-BMA Ccdc42 Glut, or
        120 MEA Otp Foxp2 Glut from 121 MEA-BST Otp Zic2 Glut.

AFTER:  A06 output is merged in per SUBCLASS, adding
          subclass_UNIQUE_marker_genes      genes passing for this anchor only
          how_to_separate_from_siblings     explicit A-positive / B-negative recipe
          GPCR_cell_type_specific           GPCRs that carry cell-type information
          GPCR_universal                    GPCRs present in most of the panel
        and three new sheets.

Adds sheets:
  Subclass_Markers_Corrected   one row per anchor, corrected markers + GPCR tiers
  Pairwise_Separators          gene ON in subclass A / OFF in subclass B
  GPCR_Specificity_Tiers       per region: which GPCRs are cell-type-specific

Run after D05.
"""
from __future__ import annotations

import argparse
import gc
import shutil
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
UNIQUE_FILL = PatternFill("solid", fgColor="C6E0B4")   # green: the actual fix
SEP_FILL = PatternFill("solid", fgColor="FFE699")      # amber: separation recipe
SPEC_FILL = PatternFill("solid", fgColor="BDD7EE")     # blue: cell-type-specific GPCR
UNIV_FILL = PatternFill("solid", fgColor="E7E6E6")     # grey: universal GPCR
WARN_FONT = Font(bold=True, color="C00000", size=10)

CORRECTED_COLUMNS = [
    "region_user",
    "cell_type_label",
    "allen_subclass_anchor",
    "role",
    "n_cells",
    "UNIQUE_separator_genes",
    "BEST_separators_also_clean_in_whole_ROI",
    "allen_name_embedded_markers_verified",
    "needs_marker_combination",
    "marker_combination_recipe",
    "how_to_separate_from_siblings",
    "roi_contamination_warning",
    "shared_marker_genes",
    "GPCRs_enriched_in_this_subclass",
    "cell_type_specific_GPCRs",
    "intermediate_GPCRs",
    "universal_GPCRs",
    "discriminating_markers_with_stats",
]

COL_WIDTHS = {
    "region_user": 9,
    "cell_type_label": 34,
    "allen_subclass_anchor": 30,
    "role": 17,
    "n_cells": 9,
    "UNIQUE_separator_genes": 46,
    "BEST_separators_also_clean_in_whole_ROI": 40,
    "roi_contamination_warning": 60,
    "allen_name_embedded_markers_verified": 24,
    "needs_marker_combination": 13,
    "how_to_separate_from_siblings": 62,
    "marker_combination_recipe": 80,
    "shared_marker_genes": 28,
    "GPCRs_enriched_in_this_subclass": 52,
    "cell_type_specific_GPCRs": 26,
    "intermediate_GPCRs": 24,
    "universal_GPCRs": 40,
    "discriminating_markers_with_stats": 70,
    "subclass_UNIQUE_marker_genes": 46,
    "subclass_BEST_markers_ROI_clean": 40,
    "GPCR_cell_type_specific": 26,
    "GPCR_universal": 40,
}

HOW_TO_READ_ROWS = [
    ("SHEET", "WHAT IT ANSWERS", "HOW TO USE IT"),
    (
        "Subclass_Markers_Corrected",
        "Which gene is UNIQUE to each Allen subclass?",
        "One row per Allen subclass anchor. UNIQUE_separator_genes = genes that pass "
        "the threshold for THIS subclass and no other subclass in the same region's "
        "panel. These are what actually separate cell type 1 / 2 / 3. Green column.",
    ),
    (
        "Subclass_Markers_Corrected",
        "Does the gene in the Allen subclass name really work?",
        "allen_name_embedded_markers_verified lists which of Allen's own name genes "
        "(Ccdc42, Foxp2, Zic2, Skor1, ...) were confirmed against expression. If a "
        "name gene is missing here it did NOT pass - see needs_marker_combination.",
    ),
    (
        "Subclass_Markers_Corrected",
        "What if no single gene is unique?",
        "needs_marker_combination = YES means no single gene is unique to this subclass, "
        "so you must use a COMBINATION. marker_combination_recipe spells it out as "
        "POSITIVE genes plus NEGATIVE gates against every other anchor in the region. "
        "Example: 005 L5 IT CTX Glut is transcriptomically intermediate between L4/5 IT "
        "and L6 IT, so it needs Rorb+ (shared with 006) together with Cux2- / Whrn- to "
        "exclude 006, and Foxp2- / Syt6- to exclude 030 L6 CT.",
    ),
    (
        "Subclass_Markers_Corrected",
        "Is my unique marker also clean against the rest of the dissection?",
        "UNIQUE_separator_genes is unique WITHIN the region's anchor panel - this is "
        "sufficient for Xenium/spatial work, where anatomy already tells you which "
        "part of the section you are in. BEST_separators_also_clean_in_whole_ROI is a "
        "much stricter extra tier: the gene must also beat all ~40-80 neuronal "
        "subclasses the Allen dissection swept up, including populations far outside "
        "your region. It is often EMPTY, and that is expected, not a failure - e.g. "
        "Slc17a7 cleanly separates 012 MEA Slc17a7 Glut from 113 inside BMAp, but the "
        "sAMY dissection also contains Slc17a7-high 009 L2/3 IT PIR-ENTl Glut. Use "
        "the strict column when you have NO anatomical constraint (e.g. dissociated "
        "cells / FACS); use UNIQUE_separator_genes for spatial panels. "
        "roi_contamination_warning names the exact competing population.",
    ),
    (
        "Pairwise_Separators",
        "Which gene is ON in subclass A but OFF in subclass B?",
        "Ordered pairs. Filter subclass_A_positive to your target, read separator_gene "
        "with pct_A (how many target cells express it) and pct_B (how many contaminant "
        "cells express it). log2_gap is the expression margin.",
    ),
    (
        "GPCR_Specificity_Tiers",
        "Is this GPCR a cell-type label or just everywhere?",
        "gpcr_tier_panel: cell_type_specific = detected in <=30% of the region's "
        "anchor subclasses, so it carries cell-type information. universal = detected "
        "in >=70%, so it is a valid DRUG target but tells you nothing about cell "
        "identity and must be paired with a cell-type marker.",
    ),
    (
        "FINAL_Recommendations / PI_Summary",
        "Where did the old marker list go?",
        "cell_type_marker_genes is unchanged (cell-type level, from the curated "
        "template). The new subclass_UNIQUE_marker_genes column next to it is the "
        "subclass-level correction. Use both: cell-type markers to find the class, "
        "UNIQUE markers to split it.",
    ),
    (
        "NOTE",
        "Why universal GPCRs are not dropped",
        "Grm5 / Gabbr1 / Gabbr2 / Grm1 / Cnr1 / Oprm1 / Adcyap1r1 are expressed across "
        "most neuronal subclasses. They remain legitimate pharmacological targets - "
        "they simply cannot be used to identify a cell type on their own.",
    ),
]


def _style_sheet(ws, columns: list[str], highlight: dict[str, PatternFill]) -> None:
    ws.freeze_panes = "A2"
    for j, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=j)
        c.fill = highlight.get(col, HDR_FILL)
        c.font = Font(bold=True, color="000000", size=10) if col in highlight else HDR_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(col, 16)
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(columns)):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=9)


def _write_sheet(wb, name: str, df: pd.DataFrame, highlight: dict[str, PatternFill]) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    cols = list(df.columns)
    ws.append(cols)
    for _, r in df.iterrows():
        ws.append(["" if pd.isna(v) else v for v in r.tolist()])
    _style_sheet(ws, cols, highlight)


def _save_with_retry(wb, path: Path, retries: int = 8, sleep: int = 20) -> Path:
    for attempt in range(1, retries + 1):
        try:
            wb.save(path)
            gc.collect()
            return path
        except PermissionError as e:
            gc.collect()
            print(f"[WARN] {path.name} locked (attempt {attempt}/{retries}): {e}")
            if attempt == retries:
                alt = path.with_name(f"{path.stem}_new{path.suffix}")
                wb.save(alt)
                print(f"[WARN] saved to {alt} instead - close Excel and rename")
                return alt
            time.sleep(sleep)
    return path


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", required=True, help="FINAL_decision.xlsx to patch in place")
    p.add_argument("--anchor_panel_csv", required=True)
    p.add_argument("--pairwise_csv", required=True)
    p.add_argument("--gpcr_tiers_csv", required=True)
    p.add_argument("--patch_sheets", nargs="*", default=["FINAL_Recommendations", "PI_Summary"])
    p.add_argument("--backup", action="store_true")
    args = p.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        raise SystemExit(f"[ERROR] workbook not found: {wb_path}")
    if args.backup:
        bak = wb_path.with_name(f"{wb_path.stem}_before_D06{wb_path.suffix}")
        shutil.copy2(wb_path, bak)
        print(f"[INFO] backup -> {bak.name}")

    anchor = pd.read_csv(args.anchor_panel_csv).fillna("")
    pairs = pd.read_csv(args.pairwise_csv)
    tiers = pd.read_csv(args.gpcr_tiers_csv)

    corrected = anchor.reindex(columns=CORRECTED_COLUMNS).fillna("")
    corrected = corrected.sort_values(
        ["region_user", "cell_type_label", "allen_subclass_anchor"]
    ).reset_index(drop=True)

    print(f"[INFO] anchors with corrected markers: {len(corrected)}")
    need_combo = corrected[corrected["needs_marker_combination"] == "YES"]
    if not need_combo.empty:
        print("[INFO] anchors with NO single unique marker (need combination):")
        for _, r in need_combo.iterrows():
            print(f"   {r['region_user']:<6} {r['allen_subclass_anchor']}")

    wb = load_workbook(wb_path)

    _write_sheet(
        wb,
        "Subclass_Markers_Corrected",
        corrected,
        {
            "UNIQUE_separator_genes": UNIQUE_FILL,
            "BEST_separators_also_clean_in_whole_ROI": UNIQUE_FILL,
            "allen_name_embedded_markers_verified": UNIQUE_FILL,
            "how_to_separate_from_siblings": SEP_FILL,
            "roi_contamination_warning": SEP_FILL,
            "needs_marker_combination": SEP_FILL,
            "marker_combination_recipe": SEP_FILL,
            "cell_type_specific_GPCRs": SPEC_FILL,
            "GPCRs_enriched_in_this_subclass": SPEC_FILL,
            "universal_GPCRs": UNIV_FILL,
        },
    )
    _write_sheet(
        wb,
        "Pairwise_Separators",
        pairs.sort_values(
            ["region_user", "subclass_A_positive", "subclass_B_negative", "log2_gap"],
            ascending=[True, True, True, False],
        ),
        {"separator_gene": UNIQUE_FILL, "log2_gap": SEP_FILL},
    )
    _write_sheet(
        wb,
        "GPCR_Specificity_Tiers",
        tiers.sort_values(["region_user", "gpcr_tier_panel", "top_panel_specificity_log2"],
                          ascending=[True, True, False]),
        {"gpcr_tier_panel": SPEC_FILL, "gpcr_tier_roi": UNIV_FILL},
    )

    # ---- patch existing recommendation sheets with subclass-level columns
    lookup = {
        (str(r["region_user"]).strip(), str(r["allen_subclass_anchor"]).strip()): r
        for _, r in anchor.iterrows()
    }
    new_cols = [
        ("subclass_UNIQUE_marker_genes", "UNIQUE_separator_genes", UNIQUE_FILL),
        ("subclass_BEST_markers_ROI_clean", "BEST_separators_also_clean_in_whole_ROI", UNIQUE_FILL),
        ("how_to_separate_from_siblings", "how_to_separate_from_siblings", SEP_FILL),
        ("GPCR_cell_type_specific", "cell_type_specific_GPCRs", SPEC_FILL),
        ("GPCR_universal", "universal_GPCRs", UNIV_FILL),
    ]

    for sheet in args.patch_sheets:
        if sheet not in wb.sheetnames:
            print(f"[WARN] sheet {sheet} not present, skipped")
            continue
        ws = wb[sheet]
        header = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
        try:
            i_region = header.index("region") + 1
            i_sub = header.index("allen_subclass") + 1
        except ValueError:
            print(f"[WARN] {sheet} lacks region/allen_subclass columns, skipped")
            continue

        next_free = ws.max_column + 1
        n_filled = 0
        for k, (out_col, src_col, fill) in enumerate(new_cols):
            # idempotent: overwrite in place when the column already exists
            if out_col in header:
                j = header.index(out_col) + 1
            else:
                j = next_free
                next_free += 1
                header.append(out_col)
            c = ws.cell(row=1, column=j, value=out_col)
            c.fill = fill
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(out_col, 30)
            for i in range(2, ws.max_row + 1):
                region = str(ws.cell(row=i, column=i_region).value or "").strip()
                sub = str(ws.cell(row=i, column=i_sub).value or "").strip()
                rec = lookup.get((region, sub))
                cell = ws.cell(row=i, column=j)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9)
                if rec is None:
                    cell.value = "(not recomputed - run A06 for this region)"
                    cell.font = WARN_FONT
                    continue
                val = rec.get(src_col, "")
                cell.value = "" if pd.isna(val) else val
                if k == 0 and str(val).strip():
                    n_filled += 1
        print(f"[OK] patched {sheet}: {n_filled}/{ws.max_row - 1} rows got unique markers")

    # ---- refresh HOW_TO_READ
    marker = "=== D06: SUBCLASS-LEVEL MARKER CORRECTION ==="
    if "HOW_TO_READ" in wb.sheetnames:
        ws = wb["HOW_TO_READ"]
        # idempotent: rewrite the D06 block in place if it is already there
        existing = [
            i for i in range(1, ws.max_row + 1) if str(ws.cell(row=i, column=1).value or "") == marker
        ]
        start = existing[0] if existing else ws.max_row + 2
        ws.cell(row=start, column=1, value=marker).font = Font(
            bold=True, size=11, color="C00000"
        )
        for i, row in enumerate(HOW_TO_READ_ROWS, start=start + 1):
            for j, v in enumerate(row, start=1):
                c = ws.cell(row=i, column=j, value=v)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.font = Font(bold=(i == start + 1), size=10)
        for j, w in enumerate((30, 46, 96), start=1):
            ws.column_dimensions[get_column_letter(j)].width = max(
                ws.column_dimensions[get_column_letter(j)].width or 0, w
            )

    saved = _save_with_retry(wb, wb_path)
    print(f"[DONE] {saved}")


if __name__ == "__main__":
    main()
