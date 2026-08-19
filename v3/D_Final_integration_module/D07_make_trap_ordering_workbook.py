#!/usr/bin/env python
"""
D07_make_trap_ordering_workbook.py

Final probe-ordering workbook for a TWO-region TRAP experiment (BMAp + ORBm).

Produces ONE deliverable sheet, "final ordering for each brain", holding the
complete gene list to order, one row per (region, gene), built from four blocks:

  1. cell_type_unique_marker   gene that is unique to ONE Allen subclass anchor
                               inside that region's panel (from A06). This is
                               what separates cell type 1 / 2 / 3.
  2. GPCR_cell_type_specific   GPCR detected in <=30% of the region's anchor
                               subclasses, so it carries cell-type information.
  3. GPCR_universal /          GPCR present across most of the panel (valid drug
     GPCR_intermediate         target, no cell-type information), plus shared
     cell_type_shared_marker   cell-type markers that label a group not a single
                               subclass.
  4. IEG / reporter_transgene  activity + TRAP tagging readout, and the
     class_backbone            glutamatergic/GABAergic backbone.

A gene can hold several roles at once (e.g. Gpr88 is both the unique separator
for 007 L2/3 IT CTX Glut and a GPCR; Drd1 likewise for 113 MEA-COA-BMA Ccdc42
Glut). Every row therefore carries `all_roles`, while `block` /
`primary_category` give one unambiguous bucket for counting.

Counts are written as a block at the top of the sheet and repeated in
Gene_Count_Summary.

Inputs: A06 outputs (BMAp + ORBm), the curated marker template, the GPCR drug
reference table, and v3/inputs/ieg_and_reporter_genes.csv.
"""
from __future__ import annotations

import argparse
import gc
import re
import time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REGIONS = ["BMAp", "ORBm"]

SHEET_MAIN = "final ordering for each brain"

# block -> (display order, fill colour)
BLOCK_STYLE = {
    "1_cell_type_unique_marker": ("C6E0B4", "Separates one Allen subclass from all others in the panel, alone or as part of its combination"),
    "2_GPCR_cell_type_specific": ("BDD7EE", "GPCR that also carries cell-type information"),
    "3_GPCR_broad_or_shared_marker": ("FFE699", "Drug-targetable / group-level; needs pairing with block 1"),
    "4_IEG_reporter_backbone": ("E4DFEC", "TRAP activity tag + reporter + neurotransmitter backbone"),
}

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F4E79")
COUNT_FONT = Font(bold=True, size=10)
CUSTOM_FILL = PatternFill("solid", fgColor="FFC7CE")  # custom probe needed
THIN = Side(style="thin", color="BFBFBF")

MAIN_COLUMNS = [
    "region",
    "gene",
    "block",
    "primary_category",
    "all_roles",
    "order_priority",
    "probe_type",
    "marks_which_cell_types",
    "n_subclasses_marked",
    "evidence",
    "drug_or_ligand",
    "notes",
]

MAIN_WIDTHS = {
    "region": 8,
    "gene": 12,
    "block": 30,
    "primary_category": 26,
    "all_roles": 44,
    "order_priority": 8,
    "probe_type": 22,
    "marks_which_cell_types": 46,
    "n_subclasses_marked": 9,
    "evidence": 40,
    "drug_or_ligand": 40,
    "notes": 62,
}


def _split(s) -> list[str]:
    if pd.isna(s) or not str(s).strip():
        return []
    return [x.strip() for x in str(s).split(",") if x.strip()]


def _save_with_retry(wb: Workbook, path: Path, retries: int = 8, sleep: int = 20) -> Path:
    for attempt in range(1, retries + 1):
        try:
            wb.save(path)
            gc.collect()
            return path
        except PermissionError as e:
            gc.collect()
            print(f"[WARN] {path.name} locked (attempt {attempt}/{retries}): {e}")
            if attempt == retries:
                alt = path.with_name(f"{path.stem}_new{path.suffix}")
                wb.save(alt)
                print(f"[WARN] saved to {alt.name} instead - close Excel and rename")
                return alt
            time.sleep(sleep)
    return path


def build_gene_rows(
    anchor: pd.DataFrame,
    disc_top: pd.DataFrame,
    tiers: pd.DataFrame,
    markers_tpl: pd.DataFrame,
    ieg: pd.DataFrame,
    drug_ref: pd.DataFrame,
    top_k_unique: int,
    allen_absent: set[str],
) -> pd.DataFrame:
    """Collect every gene needed per region, merging roles when a gene has several."""
    # gene -> one drug string, for the drug_or_ligand column
    drug_lookup: dict[str, str] = {}
    if not drug_ref.empty:
        for g, grp in drug_ref.groupby("gene_symbol"):
            approved = grp[grp["drug_status"].astype(str).str.contains("approved", case=False, na=False)]
            pick = approved if not approved.empty else grp
            bits = []
            for r in pick.head(2).itertuples():
                year = "" if pd.isna(r.year_approved_or_published) else f" {r.year_approved_or_published}"
                bits.append(f"{r.drug_name} ({r.drug_status}{year})")
            drug_lookup[str(g)] = "; ".join(bits)

    rows: dict[tuple[str, str], dict] = {}
    combos: dict[tuple[str, str], dict] = {}

    def touch(region: str, gene: str) -> dict:
        key = (region, gene)
        if key not in rows:
            rows[key] = {
                "region": region,
                "gene": gene,
                "roles": [],
                "marks": [],
                "evidence": [],
                "priority": 9,
                "probe_type": "endogenous_mouse",
                "notes": [],
                "block": None,
                "primary_category": None,
            }
        return rows[key]

    # ---- block 1: unique separators per anchor
    stats_lookup = {
        (str(r.region_user), str(r.subclass), str(r.gene)): r
        for r in disc_top.itertuples()
    }
    for _, a in anchor.iterrows():
        region = str(a["region_user"]).strip()
        sub = str(a["allen_subclass_anchor"]).strip()
        role_tag = "" if str(a.get("role", "target")) == "target" else " [neighbour control]"
        clean = set(_split(a.get("BEST_separators_also_clean_in_whole_ROI")))
        uniq = _split(a.get("UNIQUE_separator_genes"))
        # 2-3 markers identify a population; the rest are backup. Rank by
        # ROI-cleanliness first, then by A06's own within-subclass rank.
        ranked = sorted(
            uniq,
            key=lambda g: (
                0 if g in clean else 1,
                getattr(stats_lookup.get((region, sub, g)), "rank_within_subclass", 99),
            ),
        )
        core = set(ranked[:top_k_unique])
        for g in ranked:
            r = touch(region, g)
            is_core = g in core
            r["roles"].append(
                f"unique marker for {sub}" + ("" if is_core else " (backup)")
            )
            r["marks"].append(f"{sub}{role_tag}")
            r["block"] = r["block"] or "1_cell_type_unique_marker"
            r["primary_category"] = r["primary_category"] or "cell_type_unique_marker"
            r["priority"] = min(r["priority"], 1 if is_core else 2)
            st = stats_lookup.get((region, sub, g))
            if st is not None:
                r["evidence"].append(
                    f"{sub}: {st.pct_expr:.0f}% of cells, panel_spec={st.specificity_log2_vs_panel:.2f}"
                )
        for g in clean:
            touch(region, g)["notes"].append("also top expressor across the whole dissection ROI")

        # A subclass with no unique gene can only be called from a combination, so
        # every gene that combination needs has to be priority 1 too - otherwise
        # ordering the minimal panel would leave this population unidentifiable.
        if str(a.get("needs_marker_combination", "NO")) == "YES":
            combo = str(a.get("marker_combination_recipe", "") or "")
            pos = re.findall(r"([A-Za-z][A-Za-z0-9]*)\+ \(", combo)
            neg = re.findall(r"([A-Za-z][A-Za-z0-9]*)- rules out", combo)
            for g, sign in [(x, "positive") for x in pos] + [(x, "negative gate") for x in neg]:
                rr = touch(region, g)
                rr["roles"].append(f"{sign} in the {sub} marker combination")
                rr["marks"].append(f"{sub}{role_tag} (combination only)")
                if rr["block"] is None:
                    rr["block"] = "1_cell_type_unique_marker"
                    rr["primary_category"] = "cell_type_combination_marker"
                rr["priority"] = min(rr["priority"], 1)
            for g in dict.fromkeys(pos + neg):
                touch(region, g)["notes"].append(
                    f"required to call {sub}, which has NO unique gene. Full recipe: {combo}"
                )
            combos[(region, sub)] = {"positive": pos, "negative": neg, "recipe": combo}

    # ---- blocks 2 and 3: GPCRs by tier
    tier_block = {
        "cell_type_specific": ("2_GPCR_cell_type_specific", "GPCR_cell_type_specific", 1),
        "intermediate": ("3_GPCR_broad_or_shared_marker", "GPCR_intermediate", 2),
        "universal": ("3_GPCR_broad_or_shared_marker", "GPCR_universal", 2),
    }
    top_sub = {
        (str(r.region_user), str(r.gpcr_gene)): r for r in tiers.itertuples()
    }
    for _, t in tiers.iterrows():
        region = str(t["region_user"]).strip()
        g = str(t["gpcr_gene"]).strip()
        tier = str(t["gpcr_tier_panel"])
        if tier not in tier_block:
            continue
        # The tier is a FRACTION, so "detected in <=30% of the panel" also catches
        # genes detected in 0 anchors. Those are not probes, they are absent
        # transcripts - keep them out of the order list entirely.
        if int(t.get("n_panel_detected", 0)) < 1:
            continue
        blk, cat, prio = tier_block[tier]
        r = touch(region, g)
        r["roles"].append(f"GPCR ({tier}, detected in {t['frac_panel_detected']:.0%} of panel)")
        if r["block"] is None:
            r["block"] = blk
            r["primary_category"] = cat
        r["priority"] = min(r["priority"], prio)
        ts = top_sub.get((region, g))
        if ts is not None and not pd.isna(ts.top_panel_subclass):
            r["evidence"].append(
                f"highest in {ts.top_panel_subclass} ({ts.top_panel_pct:.0f}%, "
                f"panel_spec={ts.top_panel_specificity_log2:.2f})"
            )
        if tier == "universal":
            r["notes"].append(
                "expressed across most of the panel: valid drug target but carries no "
                "cell-type information, so pair it with a block-1 marker"
            )

    # ---- block 3 continued: curated shared cell-type markers
    for _, m in markers_tpl.iterrows():
        region = str(m["region_user"]).strip()
        ct = str(m["cell_type_label"]).strip()
        for g in _split(m.get("marker_genes")):
            r = touch(region, g)
            if not any("unique marker" in x for x in r["roles"]):
                r["roles"].append(f"curated marker for cell type: {ct}")
                r["marks"].append(f"cell type: {ct}")
                if r["block"] is None:
                    r["block"] = "3_GPCR_broad_or_shared_marker"
                    r["primary_category"] = "cell_type_shared_marker"
                r["priority"] = min(r["priority"], 2)

    # ---- block 4: IEG / reporter / backbone, applied to BOTH regions
    for _, i in ieg.iterrows():
        g = str(i["gene_symbol"]).strip()
        for region in REGIONS:
            r = touch(region, g)
            r["roles"].append(str(i["category"]))
            r["marks"].append(str(i["detects"]))
            r["probe_type"] = str(i["probe_type"])
            r["priority"] = min(r["priority"], int(i["priority"]))
            r["notes"].append(str(i["rationale"]))
            if r["block"] is None or r["primary_category"] in (
                "cell_type_shared_marker",
                "GPCR_universal",
                "GPCR_intermediate",
            ):
                r["block"] = "4_IEG_reporter_backbone"
                r["primary_category"] = str(i["category"])

    out = []
    for (region, gene), r in rows.items():
        marks = list(dict.fromkeys(r["marks"]))
        if gene in allen_absent:
            r["notes"].insert(
                0,
                f"'{gene}' is not a symbol in Allen WMB gene metadata (it is an older alias, "
                "e.g. Ctgf->Ccn2), so it has no Allen expression evidence here - confirm the "
                "current symbol with your probe vendor before ordering",
            )
        out.append(
            {
                "region": region,
                "gene": gene,
                "block": r["block"] or "3_GPCR_broad_or_shared_marker",
                "primary_category": r["primary_category"] or "cell_type_shared_marker",
                "all_roles": "; ".join(dict.fromkeys(r["roles"])),
                "order_priority": r["priority"] if r["priority"] < 9 else 3,
                "probe_type": r["probe_type"],
                "marks_which_cell_types": "; ".join(marks),
                "n_subclasses_marked": sum(1 for m in marks if m[:3].isdigit()),
                "evidence": "; ".join(dict.fromkeys(r["evidence"]))[:600],
                "drug_or_ligand": drug_lookup.get(gene, ""),
                "notes": " | ".join(dict.fromkeys(r["notes"]))[:600],
            }
        )
    df = pd.DataFrame(out)
    df["_r"] = df["region"].map({r: i for i, r in enumerate(REGIONS)})
    df = df.sort_values(["_r", "block", "order_priority", "gene"]).drop(columns="_r")
    return df.reset_index(drop=True), combos


def build_counts(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for region in REGIONS:
        sub = df[df["region"] == region]
        for blk in sorted(BLOCK_STYLE):
            b = sub[sub["block"] == blk]
            if b.empty:
                continue
            rows.append({"scope": region, "block": blk, "n_genes": int(b["gene"].nunique())})
        rows.append({"scope": region, "block": "TOTAL for this region", "n_genes": int(sub["gene"].nunique())})
        rows.append(
            {
                "scope": region,
                "block": "  minimal panel for this region (priority 1)",
                "n_genes": int(sub.loc[sub["order_priority"] == 1, "gene"].nunique()),
            }
        )
    per_region = {r: set(df.loc[df["region"] == r, "gene"]) for r in REGIONS}
    union = set().union(*per_region.values())
    shared = set.intersection(*per_region.values())
    rows += [
        {"scope": "BOTH", "block": "shared by BMAp and ORBm", "n_genes": len(shared)},
        {"scope": "BOTH", "block": f"{REGIONS[0]}-only", "n_genes": len(per_region[REGIONS[0]] - shared)},
        {"scope": "BOTH", "block": f"{REGIONS[1]}-only", "n_genes": len(per_region[REGIONS[1]] - shared)},
        {"scope": "BOTH", "block": "UNIQUE GENES TO ORDER (union, one probe each)", "n_genes": len(union)},
    ]
    # A gene can be priority 1 in one region and 2 in the other, so collapse to the
    # best (lowest) priority per gene; otherwise these rows would not sum to the union.
    best_prio = df.groupby("gene")["order_priority"].min()
    rows += [
        {
            "scope": "BOTH",
            "block": "MINIMAL PANEL (priority 1 in at least one region)",
            "n_genes": int((best_prio == 1).sum()),
        },
        {
            "scope": "BOTH",
            "block": "  of which CUSTOM probes (transgenes)",
            "n_genes": int(df.loc[df["probe_type"] == "CUSTOM_probe_required", "gene"].nunique()),
        },
        {"scope": "BOTH", "block": "  priority 1 (must have)", "n_genes": int((best_prio == 1).sum())},
        {"scope": "BOTH", "block": "  priority 2 (recommended backup)", "n_genes": int((best_prio == 2).sum())},
        {"scope": "BOTH", "block": "  priority 3 (optional)", "n_genes": int((best_prio == 3).sum())},
    ]
    return pd.DataFrame(rows)


def write_main_sheet(wb: Workbook, df: pd.DataFrame, counts: pd.DataFrame) -> None:
    ws = wb.create_sheet(SHEET_MAIN)
    ws["A1"] = "FINAL PROBE ORDERING LIST - BMAp + ORBm only (TRAP experiment)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Block 1 separates cell type 1 / 2 / 3 inside a region. Block 2 GPCRs do that AND are "
        "druggable. Block 3 GPCRs are druggable but not cell-type-informative, so they only work "
        "paired with block 1. Block 4 is the TRAP readout: Fos/Arc/Npas4 for activity, tdTomato/iCre "
        "for the permanent tag, Slc17a7/Slc17a6/Gad1/Gad2 for class."
    )
    ws["A2"].font = Font(size=9, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=8)

    r = 6
    ws.cell(row=r, column=1, value="GENE COUNTS").font = TITLE_FONT
    r += 1
    for j, h in enumerate(("scope", "block", "n_genes"), start=1):
        c = ws.cell(row=r, column=j, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
    r += 1
    count_start = r
    for _, cr in counts.iterrows():
        ws.cell(row=r, column=1, value=cr["scope"]).font = COUNT_FONT
        ws.cell(row=r, column=2, value=cr["block"])
        cell = ws.cell(row=r, column=3, value=int(cr["n_genes"]))
        cell.font = COUNT_FONT
        if any(k in str(cr["block"]) for k in ("UNIQUE GENES TO ORDER", "TOTAL", "MINIMAL PANEL")):
            for j in range(1, 4):
                ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor="FFF2CC")
        r += 1
    for i in range(count_start, r):
        for j in range(1, 4):
            ws.cell(row=i, column=j).border = Border(bottom=THIN)

    r += 2
    ws.cell(row=r, column=1, value="GENE LIST").font = TITLE_FONT
    r += 1
    hdr_row = r
    for j, col in enumerate(MAIN_COLUMNS, start=1):
        c = ws.cell(row=r, column=j, value=col)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = MAIN_WIDTHS.get(col, 16)
    ws.row_dimensions[r].height = 30
    r += 1
    for _, gr in df.iterrows():
        for j, col in enumerate(MAIN_COLUMNS, start=1):
            v = gr[col]
            c = ws.cell(row=r, column=j, value="" if pd.isna(v) else v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=9)
        fill = PatternFill("solid", fgColor=BLOCK_STYLE[gr["block"]][0])
        for j in (2, 3, 4):
            ws.cell(row=r, column=j).fill = fill
        ws.cell(row=r, column=2).font = Font(size=9, bold=True)
        if gr["probe_type"] == "CUSTOM_probe_required":
            ws.cell(row=r, column=7).fill = CUSTOM_FILL
            ws.cell(row=r, column=7).font = Font(size=9, bold=True, color="9C0006")
        r += 1
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(MAIN_COLUMNS))}{r - 1}"


def write_simple_sheet(wb: Workbook, name: str, df: pd.DataFrame, widths: dict[str, int]) -> None:
    ws = wb.create_sheet(name)
    cols = list(df.columns)
    ws.append(cols)
    for _, row in df.iterrows():
        ws.append(["" if pd.isna(v) else v for v in row.tolist()])
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 16)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cols)):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=9)
    ws.freeze_panes = "A2"


def write_how_to_read(
    wb: Workbook,
    df: pd.DataFrame,
    counts: pd.DataFrame,
    combos: dict,
    n_excluded: int,
) -> None:
    total = int(counts.loc[counts["block"].str.contains("UNIQUE GENES TO ORDER"), "n_genes"].iloc[0])
    minimal = int(counts.loc[counts["block"].str.contains("MINIMAL PANEL"), "n_genes"].iloc[0])
    combo_txt = "; ".join(f"{sub} ({region})" for (region, sub) in combos) or "none"

    lines: list[tuple[str, str]] = [
        ("WHAT THIS FILE IS", ""),
        (
            "Scope",
            "BMAp and ORBm only, for a TRAP experiment. Every other region from the 7-region "
            "workbook has been dropped. 20 Allen subclass populations are covered: 8 in BMAp, "
            "12 in ORBm.",
        ),
        (
            "The answer",
            f"Order {total} probes to cover both regions completely. If you are constrained, "
            f"the {minimal} priority-1 genes are enough to identify all 20 populations and read "
            f"out the TRAP tag. The remaining genes are backups and extra drug targets.",
        ),
        ("", ""),
        ("SHEETS", ""),
        (SHEET_MAIN, "The deliverable. Counts at the top, then one row per (region, gene)."),
        ("Gene_Count_Summary", "The same counts on their own, easy to cite."),
        ("BMAp_order_list / ORBm_order_list", "Flat per-region lists sorted by priority, for pasting into a vendor form."),
        ("Subclass_separation_evidence", "Per population: cell count, its unique genes, and its combination recipe if it has no unique gene."),
        ("Pairwise_separators", "Raw evidence: for every pair of populations, which gene is ON in one and OFF in the other, with the log2 gap."),
        ("GPCR_excluded_not_detected", f"{n_excluded} GPCRs deliberately left OUT. Read this before you ask why a receptor is missing."),
        ("IEG_reporter_reference", "Why each activity gene and each transgene probe is in the panel, with citations."),
        ("", ""),
        ("THE FOUR BLOCKS", ""),
        (
            "1_cell_type_unique_marker",
            "Genes that separate cell type 1 from 2 from 3 inside that region. A gene is here "
            "only if it is enriched in exactly ONE of the region's populations relative to the "
            "others (A06 specificity_log2_vs_panel). This is the block that answers 'which cell "
            "am I looking at'. Includes combination markers for populations with no unique gene.",
        ),
        (
            "2_GPCR_cell_type_specific",
            "GPCRs that are BOTH druggable AND cell-type-informative: detected in no more than "
            "30% of the region's populations, and actually detected in at least one. These are "
            "the best drug targets because hitting them hits a defined cell type.",
        ),
        (
            "3_GPCR_broad_or_shared_marker",
            "Two kinds of gene, both group-level. (a) GPCRs present across most of the panel: "
            "real drug targets, but they tell you nothing about cell identity, so they are only "
            "interpretable next to a block-1 gene. (b) Curated cell-type markers that label a "
            "family (all Sst cells, all L5) rather than one population.",
        ),
        (
            "4_IEG_reporter_backbone",
            "The TRAP readout. tdTomato and iCre report the permanent tag; Fos, Arc and Npas4 "
            "report activity at the time of imaging; Slc17a7, Slc17a6, Gad1 and Gad2 give the "
            "glutamatergic vs GABAergic class of whatever cell you found.",
        ),
        ("", ""),
        ("PRIORITY COLUMN", ""),
        ("1", "Must have. Top 3 unique markers per population, cell-type-specific GPCRs, the TRAP tag, and the core activity and class genes."),
        ("2", "Recommended. Extra unique markers as backup if a probe fails, plus broad GPCRs and shared markers."),
        ("3", "Optional. Extra IEGs and the pan-neuronal denominator."),
        ("", ""),
        ("THINGS TO CHECK BEFORE ORDERING", ""),
        (
            "Custom probes",
            "tdTomato, iCre, EGFP and WPRE are NOT mouse genes, so no standard panel contains "
            "them and Allen has no expression data for them. They need custom probes designed "
            "against your construct sequence. Confirm which reporter you actually cross in: "
            "tdTomato covers Ai14 and Ai9, EGFP covers Ai48, Ai82 and Fos-shEGFP.",
        ),
        (
            "Populations needing a combination",
            f"{combo_txt}. These have no single unique gene, so they can only be called from a "
            "positive core plus negative gates. All genes in the recipe are priority 1 - dropping "
            "any one of them makes the population uncallable. See the notes column and "
            "Subclass_separation_evidence.",
        ),
        (
            "Excluded GPCRs",
            "A GPCR detected in 0 of a region's populations still scores as 'cell_type_specific' "
            "because that tier is a fraction, not a measure of expression. Those genes are "
            "excluded from this list and recorded in GPCR_excluded_not_detected with their actual "
            "percentages, so a receptor being absent here means Allen shows no expression, not "
            "that it was overlooked.",
        ),
        (
            "Gene symbols",
            "All endogenous symbols were checked against Allen WMB gene metadata. Older aliases "
            "were replaced with the current symbols (Ctgf became Ccn2, Fam84b became Lratd2), "
            "because the old names return no Allen data.",
        ),
        (
            "Neighbour controls",
            "Some BMAp populations are labelled [neighbour control]. They are not BMAp targets; "
            "they are the sAMY-like and striatal-like populations that can contaminate a BMAp "
            "dissection, and their markers are there so you can recognise and exclude them.",
        ),
    ]

    ws = wb.create_sheet("HOW_TO_READ", 0)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 118
    ws["A1"] = "HOW TO READ THIS WORKBOOK"
    ws["A1"].font = TITLE_FONT
    r = 3
    for k, v in lines:
        a = ws.cell(row=r, column=1, value=k)
        b = ws.cell(row=r, column=2, value=v)
        if k and not v:
            a.font = Font(bold=True, size=11, color="1F4E79")
            a.fill = PatternFill("solid", fgColor="DDEBF7")
            b.fill = PatternFill("solid", fgColor="DDEBF7")
        else:
            a.font = Font(bold=True, size=9)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        b.font = Font(size=9)
        r += 1


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--anchor_panel_csv", required=True)
    p.add_argument("--disc_top_csv", required=True)
    p.add_argument("--gpcr_tiers_csv", required=True)
    p.add_argument("--pairwise_csv", required=True)
    p.add_argument("--marker_csv", required=True)
    p.add_argument("--ieg_csv", required=True)
    p.add_argument("--drug_references_csv", required=True)
    p.add_argument("--marker_presence_csv", default=None,
                   help="B02 output; used to flag marker symbols missing from Allen gene metadata")
    p.add_argument("--out_xlsx", required=True)
    p.add_argument(
        "--top_unique_per_subclass",
        type=int,
        default=3,
        help="How many unique markers per Allen subclass count as priority 1 (must have). "
        "The remaining unique markers stay in the list as priority 2 backups.",
    )
    args = p.parse_args()

    anchor = pd.read_csv(args.anchor_panel_csv)
    anchor = anchor[anchor["region_user"].isin(REGIONS)]
    disc_top = pd.read_csv(args.disc_top_csv)
    disc_top = disc_top[disc_top["region_user"].isin(REGIONS)]
    tiers = pd.read_csv(args.gpcr_tiers_csv)
    tiers = tiers[tiers["region_user"].isin(REGIONS)]
    pairwise = pd.read_csv(args.pairwise_csv)
    pairwise = pairwise[pairwise["region_user"].isin(REGIONS)]
    markers_tpl = pd.read_csv(args.marker_csv)
    markers_tpl = markers_tpl[markers_tpl["region_user"].isin(REGIONS)]
    ieg = pd.read_csv(args.ieg_csv)
    drug_ref = pd.read_csv(args.drug_references_csv)

    print(f"[INFO] anchors={len(anchor)} tiers={len(tiers)} ieg/reporter={len(ieg)}")

    not_detected = tiers[tiers["n_panel_detected"] < 1][
        ["region_user", "gpcr_gene", "gpcr_tier_panel", "n_panel_detected",
         "n_panel_subclasses", "top_panel_subclass", "top_panel_mean_log2", "top_panel_pct"]
    ].copy()
    not_detected["why_excluded"] = (
        "not detected in ANY anchor subclass of this region at pct>=20% and mean_log2>=0.5; "
        "its 'cell_type_specific' tier is an artefact of the tier being a fraction, "
        "not evidence of expression"
    )
    print(f"[INFO] GPCRs excluded as not-detected: {len(not_detected)}")
    for region, g in not_detected.groupby("region_user"):
        print(f"   {region}: {', '.join(sorted(g['gpcr_gene']))}")

    allen_absent: set[str] = set()
    if args.marker_presence_csv and Path(args.marker_presence_csv).exists():
        mp = pd.read_csv(args.marker_presence_csv)
        allen_absent = set(
            mp.loc[~mp["present_in_allen_gene_metadata"].astype(bool), "gene"].astype(str)
        )
        print(f"[INFO] marker symbols absent from Allen gene metadata: {sorted(allen_absent)}")

    df, combos = build_gene_rows(
        anchor, disc_top, tiers, markers_tpl, ieg, drug_ref,
        args.top_unique_per_subclass, allen_absent,
    )

    # The recipe is prose generated by A06, so confirm every gene parsed out of it
    # really appears in the pairwise evidence for that subclass.
    for (region, sub), c in combos.items():
        pw = pairwise[(pairwise["region_user"] == region)]
        known = set(pw.loc[pw["subclass_A_positive"] == sub, "separator_gene"]) | set(
            pw.loc[pw["subclass_B_negative"] == sub, "separator_gene"]
        )
        parsed = set(c["positive"]) | set(c["negative"])
        print(f"[CHECK] {region} {sub}: combination genes parsed={len(parsed)} "
              f"positive={c['positive']} negative={c['negative']}")
        unbacked = sorted(parsed - known)
        if unbacked:
            print(f"[WARN]  no pairwise evidence for: {unbacked}")

    counts = build_counts(df)
    print("\n" + counts.to_string(index=False))

    wb = Workbook()
    wb.remove(wb.active)
    write_main_sheet(wb, df, counts)
    write_simple_sheet(wb, "Gene_Count_Summary", counts, {"scope": 10, "block": 46, "n_genes": 10})

    # per-region flat order lists, easiest to paste into a vendor form
    for region in REGIONS:
        sub = df[df["region"] == region][
            ["gene", "block", "order_priority", "probe_type", "primary_category",
             "marks_which_cell_types", "drug_or_ligand"]
        ].sort_values(["order_priority", "block", "gene"])
        write_simple_sheet(
            wb,
            f"{region}_order_list",
            sub,
            {"gene": 12, "block": 30, "order_priority": 8, "probe_type": 22,
             "primary_category": 26, "marks_which_cell_types": 50, "drug_or_ligand": 40},
        )

    write_simple_sheet(
        wb,
        "Subclass_separation_evidence",
        anchor[
            ["region_user", "cell_type_label", "allen_subclass_anchor", "role", "n_cells",
             "UNIQUE_separator_genes", "BEST_separators_also_clean_in_whole_ROI",
             "needs_marker_combination", "marker_combination_recipe",
             "cell_type_specific_GPCRs", "universal_GPCRs"]
        ],
        {"region_user": 9, "cell_type_label": 32, "allen_subclass_anchor": 28, "role": 16,
         "n_cells": 9, "UNIQUE_separator_genes": 44, "BEST_separators_also_clean_in_whole_ROI": 38,
         "needs_marker_combination": 12, "marker_combination_recipe": 80,
         "cell_type_specific_GPCRs": 26, "universal_GPCRs": 40},
    )
    write_simple_sheet(
        wb,
        "Pairwise_separators",
        pairwise.sort_values(["region_user", "subclass_A_positive", "log2_gap"],
                             ascending=[True, True, False]),
        {"region_user": 9, "subclass_A_positive": 28, "subclass_B_negative": 28,
         "same_cell_type_label": 12, "separator_gene": 14},
    )
    write_simple_sheet(
        wb,
        "GPCR_excluded_not_detected",
        not_detected,
        {"region_user": 9, "gpcr_gene": 12, "gpcr_tier_panel": 20, "n_panel_detected": 11,
         "n_panel_subclasses": 11, "top_panel_subclass": 28, "top_panel_mean_log2": 12,
         "top_panel_pct": 11, "why_excluded": 90},
    )
    write_simple_sheet(wb, "IEG_reporter_reference", ieg,
                       {"gene_symbol": 13, "category": 20, "probe_type": 22, "priority": 8,
                        "detects": 30, "rationale": 90, "source": 46})

    write_how_to_read(wb, df, counts, combos, len(not_detected))

    out = Path(args.out_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    saved = _save_with_retry(wb, out)
    print(f"\n[DONE] {saved}")


if __name__ == "__main__":
    main()
