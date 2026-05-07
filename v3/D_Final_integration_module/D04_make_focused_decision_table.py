#!/usr/bin/env python
"""
D04_make_focused_decision_table.py

Build a single-purpose, easy-to-read decision table that COMBINES paper-curated
and Allen-WMB-10X-validated GPCRs into one decision-friendly view.

The two data sources are complementary, not competing:
- paper_suggested_gpcrs       = literature-curated genes for this cell type.
- allen_validated_top_picks   = genes that pass Allen single-cell thresholds.
- combined_GPCRs_for_probe    = UNION of the two, each gene tagged with its
                                evidence source, sorted by evidence strength:
                                    1. paper+allen_keep                  (gold)
                                    2. allen_only_keep                   (Allen-found specific)
                                    3. paper+allen_broadly_detectable    (paper supports + reliable FISH signal but not specific)
                                    4. allen_only_broadly_detectable     (Allen broadly detectable, paper silent)
                                    5. paper_only_allen_downgrade        (paper said yes, Allen has no signal)
                                Also includes inline drug info per gene (FDA / clinical / research).
- combined_evidence_summary   = compact reasoning summary of the union.
- existing_drugs_for_picks    = full drug-target details for every gene in the union.

Reads:  v3/outputs/Final_Probe_Panel_v7_modular.xlsx (sheet=Final_Summary)
Writes: v3/outputs/FINAL_decision_table_combined.xlsx (canonical)
        outputs/FINAL_decision_table_combined.xlsx    (top-level mirror)
        v3/outputs/FINAL_decision_table_combined.csv  (GitHub preview)
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


KEEP_COLUMNS = [
    "region_user",
    "cell_type_label",
    "allen_subclass_anchor",
    "n_cells_in_anchor",
    "cell_type_marker_genes",
    "exclusion_markers",
    "paper_suggested_gpcrs",
    "allen_validated_top_picks",
    "combined_GPCRs_for_probe",
    "combined_evidence_summary",
    "n_GPCRs_keep",
    "n_broadly_detectable",
    "existing_drugs_for_picks",
    "drugs_with_sources_per_picks",
    "warning",
]


def build_table(src_xlsx: Path) -> pd.DataFrame:
    df = pd.read_excel(src_xlsx, sheet_name="Final_Summary")
    cols = [c for c in KEEP_COLUMNS if c in df.columns]
    return df[cols].copy()


def write_focused_xlsx(df: pd.DataFrame, out_xlsx: Path) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    readme = pd.DataFrame(
        [
            {
                "row": 1,
                "purpose": "Single-purpose decision table that COMBINES paper-curated and Allen-WMB-10X data.",
            },
            {
                "row": 2,
                "purpose": "Each row = (region, cell type). Both data sources are used together, not against each other.",
            },
            {
                "row": 3,
                "purpose": "  paper_suggested_gpcrs       = literature-curated genes for this cell type",
            },
            {
                "row": 4,
                "purpose": "  allen_validated_top_picks   = genes that pass Allen single-cell thresholds (status=keep / candidate_to_validate)",
            },
            {
                "row": 5,
                "purpose": "  combined_GPCRs_for_probe    = UNION of paper + Allen, tagged paper+allen_keep | allen_only_keep | paper+allen_broadly_detectable | allen_only_broadly_detectable | paper_only_allen_downgrade",
            },
            {
                "row": 6,
                "purpose": "  combined_evidence_summary   = compact summary of which genes were kept and why",
            },
            {
                "row": 7,
                "purpose": "  n_GPCRs_keep                = count of cell-type-specific genes (Allen keep tier)",
            },
            {
                "row": 8,
                "purpose": "  n_broadly_detectable        = count of broadly-expressed but reliably-detectable genes (use only if region/spatial constrained)",
            },
            {
                "row": 9,
                "purpose": "  existing_drugs_for_picks    = FDA-approved + clinical/research drugs (compact list, no source IDs)",
            },
            {
                "row": 10,
                "purpose": "  drugs_with_sources_per_picks = SAME drugs but with DrugBank ID, FDA application, key PubMed PMID + URL (clickable). One header per gene.",
            },
            {
                "row": 11,
                "purpose": "Recommended order = combined_GPCRs_for_probe (paper+allen_keep > allen_only_keep > broadly_detectable > paper_only)",
            },
            {
                "row": 12,
                "purpose": "Full detailed drug-reference table is in workbook sheet 'GPCR_Drug_References' (~94 rows; one row per drug with DrugBank URL + PubMed URL).",
            },
            {
                "row": 13,
                "purpose": "Generated by D04_make_focused_decision_table.py from v3/outputs/Final_Probe_Panel_v7_modular.xlsx (Final_Summary sheet)",
            },
        ]
    )

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
        readme.to_excel(w, sheet_name="HOW_TO_READ", index=False)
        df.to_excel(w, sheet_name="Decision_Table", index=False)

    wb = load_workbook(out_xlsx)
    ws = wb["Decision_Table"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    paper_fill = PatternFill("solid", fgColor="FFD966")  # yellow
    allen_fill = PatternFill("solid", fgColor="A9D08E")  # green
    union_fill = PatternFill("solid", fgColor="9CC2E5")  # blue (the answer)
    summary_fill = PatternFill("solid", fgColor="F4B084")  # orange
    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    column_widths = {
        "region_user": 8,
        "cell_type_label": 30,
        "allen_subclass_anchor": 22,
        "n_cells_in_anchor": 8,
        "cell_type_marker_genes": 32,
        "exclusion_markers": 22,
        "paper_suggested_gpcrs": 50,
        "allen_validated_top_picks": 60,
        "combined_GPCRs_for_probe": 130,
        "combined_evidence_summary": 90,
        "n_GPCRs_keep": 10,
        "n_broadly_detectable": 12,
        "existing_drugs_for_picks": 90,
        "drugs_with_sources_per_picks": 110,
        "warning": 30,
    }

    drug_fill = PatternFill("solid", fgColor="C9A0DC")  # purple
    drug_src_fill = PatternFill("solid", fgColor="B5A0DC")  # darker purple
    headers = [c.value for c in ws[1]]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = bold_white
        cell.fill = header_fill
        if h == "paper_suggested_gpcrs":
            cell.fill = paper_fill
            cell.font = bold_black
        elif h == "allen_validated_top_picks":
            cell.fill = allen_fill
            cell.font = bold_black
        elif h == "combined_GPCRs_for_probe":
            cell.fill = union_fill
            cell.font = bold_black
        elif h == "combined_evidence_summary":
            cell.fill = summary_fill
            cell.font = bold_black
        elif h == "existing_drugs_for_picks":
            cell.fill = drug_fill
            cell.font = bold_black
        elif h == "drugs_with_sources_per_picks":
            cell.fill = drug_src_fill
            cell.font = bold_black
        ws.column_dimensions[get_column_letter(idx)].width = column_widths.get(h, 20)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 110

    ws.freeze_panes = "C2"
    wb.save(out_xlsx)


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--src_xlsx", default="v3/outputs/Final_Probe_Panel_v7_modular.xlsx")
    p.add_argument("--out_xlsx", default="v3/outputs/FINAL_decision_table_combined_with_sources.xlsx")
    p.add_argument("--top_mirror", default="outputs/FINAL_decision_table_combined_with_sources.xlsx")
    p.add_argument("--csv_preview", default="v3/outputs/FINAL_decision_table_combined_with_sources.csv")
    args = p.parse_args()

    df = build_table(Path(args.src_xlsx))
    print(f"[INFO] decision table shape: {df.shape}")
    write_focused_xlsx(df, Path(args.out_xlsx))
    print(f"[OK] {args.out_xlsx}")
    write_focused_xlsx(df, Path(args.top_mirror))
    print(f"[OK] {args.top_mirror}  (top-level mirror)")
    df.to_csv(args.csv_preview, index=False)
    print(f"[OK] {args.csv_preview}  (GitHub-renderable preview)")


if __name__ == "__main__":
    main()
