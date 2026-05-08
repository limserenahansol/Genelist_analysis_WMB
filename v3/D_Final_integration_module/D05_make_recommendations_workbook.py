#!/usr/bin/env python
"""
D05_make_recommendations_workbook.py

Build the COMBINED 6-sheet decision workbook (one file, everything you need):

  Sheet 1  HOW_TO_READ                       - legend explaining the sheets
  Sheet 2  PI_Summary                        - PI-FACING short view (10 cols)
  Sheet 3  FINAL_Recommendations             - conclusions only
  Sheet 4  Decision_Table_full               - drilldown with metrics
  Sheet 5  GPCR_Drug_References              - long-format drug citation
  Sheet 6  MarkerBased_Cluster_Recall_TopN   - cluster-level recall using
                                                paper markers (LM/RE/BMAp
                                                only, since these are the
                                                anatomically-broad ROI regions)

Reads:  v3/outputs/Final_Probe_Panel_v8_modular.xlsx (FINAL_Recommendations,
        Final_Summary, GPCR_Drug_References sheets)
        v3/outputs/marker_recall/MarkerBased_Cluster_Recall_TopN.csv (optional)
Writes: v3/outputs/FINAL_decision.xlsx (canonical, styled)
        outputs/FINAL_decision.xlsx    (top-level mirror)
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REC_COLUMN_WIDTHS = {
    "region": 8,
    "cell_type": 32,
    "allen_subclass": 28,
    "role": 18,
    "n_cells_in_anchor": 10,
    "region_dissection_warning": 60,
    "cell_type_marker_genes": 38,
    "exclusion_markers": 22,
    "recommended_GPCR_panel": 38,
    "recommended_drugs_per_gene": 70,
    "n_genes_recommended": 10,
    "evidence_tier_per_gene": 60,
    "what_to_do": 40,
}

DECISION_TABLE_COLUMNS = [
    "region_user",
    "cell_type_label",
    "allen_subclass_anchor",
    "anchor_role",
    "n_cells_in_anchor",
    "region_dissection_warning",
    "cell_type_marker_genes",
    "exclusion_markers",
    "paper_suggested_gpcrs",
    "allen_validated_top_picks",
    "combined_GPCRs_for_probe",
    "combined_evidence_summary",
    "n_GPCRs_keep",
    "n_broadly_detectable",
    "existing_drugs_for_picks",
    "drugs_with_sources_per_picks",
    "warning",
]

DECISION_COLUMN_WIDTHS = {
    "region_user": 8,
    "cell_type_label": 30,
    "allen_subclass_anchor": 22,
    "anchor_role": 18,
    "n_cells_in_anchor": 8,
    "region_dissection_warning": 60,
    "cell_type_marker_genes": 32,
    "exclusion_markers": 22,
    "paper_suggested_gpcrs": 50,
    "allen_validated_top_picks": 60,
    "combined_GPCRs_for_probe": 130,
    "combined_evidence_summary": 90,
    "n_GPCRs_keep": 10,
    "n_broadly_detectable": 12,
    "existing_drugs_for_picks": 90,
    "drugs_with_sources_per_picks": 110,
    "warning": 30,
}


def _style_recommendations_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    marker_fill = PatternFill("solid", fgColor="F4B084")  # orange
    gene_fill = PatternFill("solid", fgColor="9CC2E5")  # blue
    drug_fill = PatternFill("solid", fgColor="C9A0DC")  # purple
    action_fill = PatternFill("solid", fgColor="A9D08E")  # green
    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [c.value for c in ws[1]]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = bold_white
        cell.fill = header_fill
        if h == "cell_type_marker_genes":
            cell.fill = marker_fill
            cell.font = bold_black
        elif h == "recommended_GPCR_panel":
            cell.fill = gene_fill
            cell.font = bold_black
        elif h == "recommended_drugs_per_gene":
            cell.fill = drug_fill
            cell.font = bold_black
        elif h == "what_to_do":
            cell.fill = action_fill
            cell.font = bold_black
        ws.column_dimensions[get_column_letter(idx)].width = REC_COLUMN_WIDTHS.get(h, 18)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    # row coloring by what_to_do
    if "what_to_do" in headers:
        action_idx = headers.index("what_to_do") + 1
        order_now_fill = PatternFill("solid", fgColor="E2EFDA")
        spatial_fill = PatternFill("solid", fgColor="FFF2CC")
        review_fill = PatternFill("solid", fgColor="FCE4D6")
        neighbor_fill = PatternFill("solid", fgColor="D9D9D9")
        for r in range(2, ws.max_row + 1):
            v = str(ws.cell(row=r, column=action_idx).value or "")
            row_fill = None
            if v.startswith("ORDER ("):
                row_fill = order_now_fill
            elif v.startswith("ORDER WITH SPATIAL"):
                row_fill = spatial_fill
            elif v.startswith("NEIGHBOR CONTROL") or v.startswith("EXCLUSION"):
                row_fill = neighbor_fill
            elif v.startswith("REVIEW"):
                row_fill = review_fill
            if row_fill is not None:
                for c in range(1, ws.max_column + 1):
                    target = ws.cell(row=r, column=c)
                    if target.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                        target.fill = row_fill

    ws.row_dimensions[1].height = 32
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 60
    ws.freeze_panes = "C2"


def _style_decision_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    paper_fill = PatternFill("solid", fgColor="FFD966")
    allen_fill = PatternFill("solid", fgColor="A9D08E")
    union_fill = PatternFill("solid", fgColor="9CC2E5")
    summary_fill = PatternFill("solid", fgColor="F4B084")
    drug_fill = PatternFill("solid", fgColor="C9A0DC")
    drug_src_fill = PatternFill("solid", fgColor="B5A0DC")
    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [c.value for c in ws[1]]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = bold_white
        cell.fill = header_fill
        if h == "paper_suggested_gpcrs":
            cell.fill = paper_fill
            cell.font = bold_black
        elif h == "allen_validated_top_picks":
            cell.fill = allen_fill
            cell.font = bold_black
        elif h == "combined_GPCRs_for_probe":
            cell.fill = union_fill
            cell.font = bold_black
        elif h == "combined_evidence_summary":
            cell.fill = summary_fill
            cell.font = bold_black
        elif h == "existing_drugs_for_picks":
            cell.fill = drug_fill
            cell.font = bold_black
        elif h == "drugs_with_sources_per_picks":
            cell.fill = drug_src_fill
            cell.font = bold_black
        ws.column_dimensions[get_column_letter(idx)].width = DECISION_COLUMN_WIDTHS.get(h, 20)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 110
    ws.freeze_panes = "C2"


def _style_drug_refs_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    bold_white = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    widths = {
        "gene_symbol": 12,
        "iuphar_receptor": 30,
        "drug_name": 32,
        "drug_status": 18,
        "drug_class": 25,
        "year_approved_or_published": 22,
        "indication": 38,
        "drugbank_id": 14,
        "drugbank_url": 42,
        "fda_application": 16,
        "key_pmid": 14,
        "pubmed_url": 42,
        "notes": 50,
    }
    headers = [c.value for c in ws[1]]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = bold_white
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 18)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"


PI_SUMMARY_COLUMNS = [
    "region",
    "cell_type",
    "allen_subclass",
    "cell_type_marker_genes",
    "exclusion_markers",
    "recommended_GPCR_panel",
    "recommended_drugs_per_gene",
    "n_genes_recommended",
    "evidence_tier_per_gene",
    "what_to_do",
]

PI_SUMMARY_WIDTHS = {
    "region": 9,
    "cell_type": 34,
    "allen_subclass": 30,
    "cell_type_marker_genes": 38,
    "exclusion_markers": 26,
    "recommended_GPCR_panel": 38,
    "recommended_drugs_per_gene": 70,
    "n_genes_recommended": 8,
    "evidence_tier_per_gene": 50,
    "what_to_do": 38,
}


def _style_pi_summary_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    marker_fill = PatternFill("solid", fgColor="F4B084")  # orange (cell-type markers)
    excl_fill = PatternFill("solid", fgColor="D9D9D9")    # grey (exclusion markers)
    gene_fill = PatternFill("solid", fgColor="9CC2E5")    # blue (GPCRs)
    drug_fill = PatternFill("solid", fgColor="C9A0DC")    # purple (drugs)
    action_fill = PatternFill("solid", fgColor="A9D08E")  # green (what_to_do)
    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [c.value for c in ws[1]]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = bold_white
        cell.fill = header_fill
        if h == "cell_type_marker_genes":
            cell.fill = marker_fill
            cell.font = bold_black
        elif h == "exclusion_markers":
            cell.fill = excl_fill
            cell.font = bold_black
        elif h == "recommended_GPCR_panel":
            cell.fill = gene_fill
            cell.font = bold_black
        elif h == "recommended_drugs_per_gene":
            cell.fill = drug_fill
            cell.font = bold_black
        elif h == "what_to_do":
            cell.fill = action_fill
            cell.font = bold_black
        ws.column_dimensions[get_column_letter(idx)].width = PI_SUMMARY_WIDTHS.get(h, 18)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    if "what_to_do" in headers:
        action_idx = headers.index("what_to_do") + 1
        order_now_fill = PatternFill("solid", fgColor="E2EFDA")
        spatial_fill = PatternFill("solid", fgColor="FFF2CC")
        review_fill = PatternFill("solid", fgColor="FCE4D6")
        neighbor_fill = PatternFill("solid", fgColor="D9D9D9")
        for r in range(2, ws.max_row + 1):
            v = str(ws.cell(row=r, column=action_idx).value or "")
            row_fill = None
            if v.startswith("ORDER ("):
                row_fill = order_now_fill
            elif v.startswith("ORDER WITH SPATIAL"):
                row_fill = spatial_fill
            elif v.startswith("NEIGHBOR CONTROL") or v.startswith("EXCLUSION"):
                row_fill = neighbor_fill
            elif v.startswith("REVIEW"):
                row_fill = review_fill
            if row_fill is not None:
                for c in range(1, ws.max_column + 1):
                    target = ws.cell(row=r, column=c)
                    if target.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                        target.fill = row_fill

    ws.row_dimensions[1].height = 32
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 60
    ws.freeze_panes = "C2"


def _style_marker_recall_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    target_fill = PatternFill("solid", fgColor="C6E0B4")
    neighbor_fill = PatternFill("solid", fgColor="D9D9D9")
    bold_white = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    widths = {
        "region_user": 8,
        "cell_type_label": 32,
        "role": 18,
        "cluster": 28,
        "subclass": 30,
        "supertype": 28,
        "n_cells_in_cluster": 10,
        "n_pos_markers_total": 10,
        "n_pos_markers_in_data": 10,
        "n_neg_markers_in_data": 10,
        "mean_pos_log2": 11,
        "mean_neg_log2": 11,
        "pct_pos_avg": 10,
        "pct_neg_avg": 10,
        "n_pos_above_30pct": 12,
        "frac_pos_above_30pct": 12,
        "net_score": 11,
        "specificity": 11,
        "rank_within_celltype": 10,
    }
    headers = [c.value for c in ws[1]]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = bold_white
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 14)
    role_idx = headers.index("role") + 1 if "role" in headers else None
    for r in range(2, ws.max_row + 1):
        if role_idx:
            v = str(ws.cell(row=r, column=role_idx).value or "")
            row_fill = target_fill if v == "target" else (neighbor_fill if v == "neighbor_control" else None)
            if row_fill is not None:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = row_fill
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = wrap
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "D2"


def _make_legend() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"sheet": "HOW_TO_READ", "row": 1, "purpose": "This legend. Explains the four sheets and how to use them."},
            {"sheet": "", "row": 2, "purpose": ""},
            {"sheet": "FINAL_Recommendations", "row": 3, "purpose": "ONE row per (region, cell type). The conclusion: which markers to use, which GPCRs to order probes for, which drug per gene. No metrics, no URLs."},
            {"sheet": "FINAL_Recommendations", "row": 4, "purpose": "Key columns: cell_type_marker_genes (orange) | recommended_GPCR_panel (blue) | recommended_drugs_per_gene (purple, marked '(FDA YEAR)' / '(clinical-trial)' / '(research)') | what_to_do (green ORDER / yellow SPATIAL CONSTRAINT / orange REVIEW)."},
            {"sheet": "FINAL_Recommendations", "row": 5, "purpose": "Tier priority for picks: paper+allen_keep > allen_only_keep > paper+allen_broadly_detectable > allen_only_broadly_detectable. Up to 5 GPCRs per cell type."},
            {"sheet": "", "row": 6, "purpose": ""},
            {"sheet": "Decision_Table_full", "row": 7, "purpose": "Same 18 rows but with full reasoning: paper_suggested vs Allen-validated, combined_GPCRs_for_probe with all metrics, drugs_with_sources_per_picks with DrugBank IDs + FDA NDA + PubMed PMIDs."},
            {"sheet": "Decision_Table_full", "row": 8, "purpose": "Read this sheet when you need to justify a pick or look up DrugBank / PubMed for a particular drug."},
            {"sheet": "", "row": 9, "purpose": ""},
            {"sheet": "GPCR_Drug_References", "row": 10, "purpose": "Long-format drug citation table - 94 rows, ONE row per (gene, drug). Columns: drug_status, year, indication, drugbank_id, drugbank_url, fda_application, key_pmid, pubmed_url."},
            {"sheet": "GPCR_Drug_References", "row": 11, "purpose": "Use this sheet to look up additional drugs for any GPCR (FINAL_Recommendations only shows the top drug per gene)."},
            {"sheet": "", "row": 12, "purpose": ""},
            {"sheet": "MarkerBased_Cluster_Recall", "row": 13, "purpose": "NEW: Paper-marker-based recall of Allen WMB clusters for LM, RE, BMAp (the 3 regions with broad ROI dissection). For each cell type, the top 5 Allen clusters that best match the paper marker signature are listed with net_score, pct_pos_avg, n_pos_markers_above_30pct, etc."},
            {"sheet": "MarkerBased_Cluster_Recall", "row": 14, "purpose": "Compare against the subclass anchor in CellType_Subclass_Anchors. If the recall returns a cluster from a DIFFERENT subclass than the v6 anchor, the paper marker set may be either (a) imprecise for the intended anatomical target, or (b) actually identifying a transcriptomically-related neighboring population. Inspect manually."},
            {"sheet": "", "row": 15, "purpose": ""},
            {"sheet": "PI_Summary", "row": 16, "purpose": "PI-FACING short view: ONE row per (region, cell type, subclass) with only 10 columns (region | cell_type | allen_subclass | cell_type_marker_genes | exclusion_markers | recommended_GPCR_panel | recommended_drugs_per_gene | n_genes_recommended | evidence_tier_per_gene | what_to_do). No technical metadata, no warnings, no DrugBank IDs. Use this when sharing with PI."},
            {"sheet": "", "row": 17, "purpose": ""},
            {"sheet": "All sheets", "row": 18, "purpose": "Generated from v3/outputs/Final_Probe_Panel_v8_modular.xlsx by D05_make_recommendations_workbook.py. Cluster recall computed by A05_marker_based_cluster_recall.py."},
        ]
    )


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--src_xlsx", default="v3/outputs/Final_Probe_Panel_v8_modular.xlsx")
    p.add_argument("--out_xlsx", default="v3/outputs/FINAL_decision.xlsx")
    p.add_argument("--top_mirror", default="outputs/FINAL_decision.xlsx")
    p.add_argument("--marker_recall_csv",
                   default="v3/outputs/marker_recall/MarkerBased_Cluster_Recall_TopN.csv")
    args = p.parse_args()

    src = Path(args.src_xlsx)
    rec = pd.read_excel(src, sheet_name="FINAL_Recommendations")
    summary = pd.read_excel(src, sheet_name="Final_Summary")
    drug_refs = pd.read_excel(src, sheet_name="GPCR_Drug_References")
    decision = summary[[c for c in DECISION_TABLE_COLUMNS if c in summary.columns]].copy()
    pi_summary = rec[[c for c in PI_SUMMARY_COLUMNS if c in rec.columns]].copy()
    legend = _make_legend()

    marker_recall_df = pd.DataFrame()
    marker_recall_path = Path(args.marker_recall_csv)
    if marker_recall_path.exists():
        marker_recall_df = pd.read_csv(marker_recall_path)
        print(f"[INFO] MarkerBased_Cluster_Recall_TopN: {marker_recall_df.shape}")

    print(
        f"[INFO] PI_Summary: {pi_summary.shape}; "
        f"FINAL_Recommendations: {rec.shape}; "
        f"Decision_Table_full: {decision.shape}; "
        f"GPCR_Drug_References: {drug_refs.shape}"
    )

    for out_path in (Path(args.out_xlsx), Path(args.top_mirror)):
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(out_path, engine="openpyxl") as w:
            legend.to_excel(w, sheet_name="HOW_TO_READ", index=False)
            pi_summary.to_excel(w, sheet_name="PI_Summary", index=False)
            rec.to_excel(w, sheet_name="FINAL_Recommendations", index=False)
            decision.to_excel(w, sheet_name="Decision_Table_full", index=False)
            drug_refs.to_excel(w, sheet_name="GPCR_Drug_References", index=False)
            if not marker_recall_df.empty:
                marker_recall_df.to_excel(w, sheet_name="MarkerBased_Cluster_Recall", index=False)

        wb = load_workbook(out_path)
        # Style HOW_TO_READ minimally
        ws = wb["HOW_TO_READ"]
        for col_letter, w_ in zip(["A", "B", "C"], [22, 6, 110]):
            ws.column_dimensions[col_letter].width = w_
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 24
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 32

        if "PI_Summary" in wb.sheetnames:
            _style_pi_summary_sheet(wb["PI_Summary"])
        _style_recommendations_sheet(wb["FINAL_Recommendations"])
        _style_decision_sheet(wb["Decision_Table_full"])
        _style_drug_refs_sheet(wb["GPCR_Drug_References"])
        if "MarkerBased_Cluster_Recall" in wb.sheetnames:
            _style_marker_recall_sheet(wb["MarkerBased_Cluster_Recall"])
        wb.save(out_path)
        print(f"[OK] {out_path}")


if __name__ == "__main__":
    main()
