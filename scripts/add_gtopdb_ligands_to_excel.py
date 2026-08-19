#!/usr/bin/env python
"""Fetch IUPHAR/GtoPdb ligands for GPCRs in Hansol 7-region workbook and update Excel."""
from __future__ import annotations

import json
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = "https://www.guidetopharmacology.org/services"
INPUT_XLSX = Path(r"c:\Users\hsollim\Downloads\genelist_for_hansol_7brainregions.xlsx")
OUTPUT_XLSX = INPUT_XLSX  # overwrite in place
DETAILED_CSV = Path(__file__).resolve().parents[1] / "v3/inputs/gpcr_drug_targets_detailed.csv"

# Mouse gene symbols in workbook (capitalized first letter for GtoP query)
GENES_ORDERED: list[str] = []


def _fetch_json(url: str, retries: int = 3) -> list | dict:
    req = urllib.request.Request(url, headers={"User-Agent": "GenelistAnalysis/1.0"})
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=45) as resp:
                return json.loads(resp.read())
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as e:
            last_err = e
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"Failed {url}: {last_err}")


def _clean_html(s: str) -> str:
    return re.sub(r"<[^>]+>", "", str(s or ""))


def _affinity_float(row: dict) -> float:
    try:
        return float(row.get("affinity") or 0)
    except (TypeError, ValueError):
        return 0.0


def lookup_target(gene: str) -> dict | None:
    q = urllib.parse.urlencode({"geneSymbol": gene, "type": "GPCR"})
    data = _fetch_json(f"{BASE}/targets?{q}")
    if not data:
        return None
    # Prefer exact gene match when multiple (rare)
    return data[0]


def fetch_ligands_for_target(target_id: int, limit: int = 25) -> list[dict]:
    data = _fetch_json(f"{BASE}/targets/{target_id}/interactions")
    if not isinstance(data, list):
        return []
    rows = []
    for row in data:
        if row.get("endogenous"):
            continue
        rows.append(row)
    rows.sort(key=_affinity_float, reverse=True)
    return rows[:limit]


def _gtop_ligand_url(ligand_id: int) -> str:
    return f"https://www.guidetopharmacology.org/GRAC/LigandDisplayForward?ligandId={ligand_id}"


def _gtop_target_url(target_id: int) -> str:
    return f"https://www.guidetopharmacology.org/GRAC/ObjectDisplayForward?objectId={target_id}"


def format_ligand_short(row: dict) -> str:
    name = row.get("ligandName", "")
    typ = row.get("type", "")
    action = row.get("action", "")
    aff = row.get("affinity", "")
    aff_type = row.get("affinityParameter", "")
    parts = [name]
    if typ or action:
        parts.append(f"({typ}{': ' + action if action else ''})".replace("()", ""))
    if aff and aff_type:
        parts.append(f"{aff_type}={aff}")
    return " ".join(parts).strip()


def build_gtop_tables(genes: list[str]) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, dict]]:
    long_rows: list[dict] = []
    summary_by_gene: dict[str, dict] = {}

    for gene in sorted(set(genes)):
        target = lookup_target(gene)
        if not target:
            summary_by_gene[gene] = {
                "gene_symbol": gene,
                "gtop_target_id": "",
                "iuphar_receptor_name": "",
                "gtop_target_url": "",
                "gtop_ligands_top": "",
                "n_ligands_in_gtop": 0,
            }
            continue

        tid = target["targetId"]
        tname = _clean_html(target.get("name", ""))
        interactions = fetch_ligands_for_target(tid)
        top_strs = []
        for i, row in enumerate(interactions[:8], start=1):
            lid = row.get("ligandId")
            long_rows.append(
                {
                    "gene_symbol": gene,
                    "iuphar_receptor_name": tname,
                    "gtop_target_id": tid,
                    "gtop_target_url": _gtop_target_url(tid),
                    "ligand_name": row.get("ligandName", ""),
                    "interaction_type": row.get("type", ""),
                    "action": row.get("action", ""),
                    "affinity": row.get("affinity", ""),
                    "affinity_parameter": row.get("affinityParameter", ""),
                    "target_species": row.get("targetSpecies", ""),
                    "endogenous": row.get("endogenous", False),
                    "gtop_ligand_id": lid,
                    "gtop_ligand_url": _gtop_ligand_url(lid) if lid else "",
                    "rank_by_affinity": i,
                }
            )
            top_strs.append(format_ligand_short(row))

        summary_by_gene[gene] = {
            "gene_symbol": gene,
            "gtop_target_id": tid,
            "iuphar_receptor_name": tname,
            "gtop_target_url": _gtop_target_url(tid),
            "gtop_ligands_top": " | ".join(top_strs),
            "n_ligands_in_gtop": len(interactions),
        }
        time.sleep(0.15)

    long_df = pd.DataFrame(long_rows)
    summary_df = pd.DataFrame(list(summary_by_gene.values()))
    return long_df, summary_df, summary_by_gene


def expand_panel_ligands(panel: str, summary_by_gene: dict[str, dict]) -> str:
    if pd.isna(panel) or not str(panel).strip():
        return ""
    out = []
    for g in [x.strip() for x in str(panel).split(",") if x.strip()]:
        info = summary_by_gene.get(g, {})
        lig = info.get("gtop_ligands_top", "")
        if lig:
            out.append(f"{g}: {lig.split(' | ')[0]}")  # top-1 per gene
        else:
            out.append(f"{g}: (no GtoPdb ligand)")
    return " | ".join(out)


def main() -> None:
    if not INPUT_XLSX.exists():
        raise SystemExit(f"Missing {INPUT_XLSX}")

    xl = pd.ExcelFile(INPUT_XLSX)
    sheets: dict[str, pd.DataFrame] = {s: pd.read_excel(INPUT_XLSX, sheet_name=s) for s in xl.sheet_names}

    final = sheets["FINAL_SUMARY"]
    genes: set[str] = set()
    for v in final["recommended_GPCR_panel"].dropna():
        genes.update(g.strip() for g in str(v).split(",") if g.strip())

    print(f"[INFO] Fetching GtoPdb ligands for {len(genes)} genes...")
    long_df, summary_df, summary_by_gene = build_gtop_tables(sorted(genes))
    print(f"[OK] GtoP long table: {long_df.shape}; summary: {summary_df.shape}")

    # Merge FDA/clinical from existing detailed CSV
    if DETAILED_CSV.exists():
        det = pd.read_csv(DETAILED_CSV)
        det_top = (
            det.sort_values(["gene_symbol", "drug_status"])
            .groupby("gene_symbol", as_index=False)
            .first()
        )
        summary_df = summary_df.merge(
            det_top[["gene_symbol", "drug_name", "drug_status", "drugbank_url", "fda_application"]],
            on="gene_symbol",
            how="left",
        )

    # Enrich FINAL_SUMARY rows
    enriched = final.copy()
    gene_to_iuphar = summary_df.set_index("gene_symbol")["iuphar_receptor_name"].to_dict()
    gene_to_url = summary_df.set_index("gene_symbol")["gtop_target_url"].to_dict()

    def panel_iuphar(panel):
        if pd.isna(panel):
            return ""
        return "; ".join(
            f"{g}={_clean_html(gene_to_iuphar.get(g.strip(), 'NA'))}"
            for g in str(panel).split(",")
            if g.strip()
        )

    def panel_urls(panel):
        if pd.isna(panel):
            return ""
        return "; ".join(gene_to_url.get(g.strip(), "") for g in str(panel).split(",") if g.strip())

    enriched["iuphar_receptor_names"] = enriched["recommended_GPCR_panel"].map(panel_iuphar)
    enriched["GtoP_target_urls"] = enriched["recommended_GPCR_panel"].map(panel_urls)
    enriched["GtoP_top_ligands_per_gene"] = enriched["recommended_GPCR_panel"].map(
        lambda p: expand_panel_ligands(p, summary_by_gene)
    )

    # Workflow note sheet
    workflow = pd.DataFrame(
        [
            {
                "step": 1,
                "what": "Transcriptomic cell x gene matrix",
                "tool": "Allen ABC / WMB-10X or your own scRNA-seq",
                "notes": "Same starting point as colleague workflow",
            },
            {
                "step": 2,
                "what": "Cell type assignment",
                "tool": "Allen mapMyCells (or subclass anchors in this workbook)",
                "notes": "Assign Cell Type A etc.",
            },
            {
                "step": 3,
                "what": "DEG / enriched genes in cell type of interest",
                "tool": "Seurat FindMarkers (colleague) or Allen GPCR ranking (this pipeline)",
                "notes": "recommended_GPCR_panel column",
            },
            {
                "step": 4,
                "what": "Filter to targetable receptors",
                "tool": "GPCR universe list + IUPHAR family",
                "notes": "33 GPCRs in FINAL_SUMARY",
            },
            {
                "step": 5,
                "what": "Ligand lookup",
                "tool": "IUPHAR/BPS Guide to PHARMACOLOGY (guidetopharmacology.org)",
                "notes": "GtoP_Ligand_References sheet = API pull; check purchasability separately (Tocris/Sigma/MCE)",
            },
        ]
    )

    sources = sheets.get("Workbook_Sources", pd.DataFrame())
    new_src = pd.DataFrame(
        [
            {
                "Source name": "IUPHAR/BPS Guide to PHARMACOLOGY (GtoPdb)",
                "Why useful": "Curated receptor–ligand interactions; ligand type (agonist/antagonist), affinity, links to purchasable tool compounds and approved drugs.",
                "URL": "https://www.guidetopharmacology.org/",
            },
            {
                "Source name": "GtoPdb REST API",
                "Why useful": "Programmatic ligand fetch: /services/targets?geneSymbol=GENE&type=GPCR then /targets/{id}/interactions",
                "URL": "https://www.guidetopharmacology.org/webServices.jsp",
            },
        ]
    )
    if not sources.empty and "Source name" in sources.columns:
        sources = pd.concat([sources, new_src], ignore_index=True).drop_duplicates(subset=["Source name"])
    else:
        sources = new_src

    out_path = OUTPUT_XLSX
    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as w:
            enriched.to_excel(w, sheet_name="FINAL_SUMARY", index=False)
            sheets.get("draft", pd.DataFrame()).to_excel(w, sheet_name="draft", index=False)
            sources.to_excel(w, sheet_name="Workbook_Sources", index=False)
            workflow.to_excel(w, sheet_name="Colleague_Workflow", index=False)
            long_df.to_excel(w, sheet_name="GtoP_Ligand_References", index=False)
            summary_df.to_excel(w, sheet_name="GtoP_Receptor_Summary", index=False)
            if DETAILED_CSV.exists():
                pd.read_csv(DETAILED_CSV).to_excel(w, sheet_name="FDA_Drug_References", index=False)
    except PermissionError:
        out_path = INPUT_XLSX.with_name(INPUT_XLSX.stem + "_with_GtoPdb.xlsx")
        with pd.ExcelWriter(out_path, engine="openpyxl") as w:
            enriched.to_excel(w, sheet_name="FINAL_SUMARY", index=False)
            sheets.get("draft", pd.DataFrame()).to_excel(w, sheet_name="draft", index=False)
            sources.to_excel(w, sheet_name="Workbook_Sources", index=False)
            workflow.to_excel(w, sheet_name="Colleague_Workflow", index=False)
            long_df.to_excel(w, sheet_name="GtoP_Ligand_References", index=False)
            summary_df.to_excel(w, sheet_name="GtoP_Receptor_Summary", index=False)
            if DETAILED_CSV.exists():
                pd.read_csv(DETAILED_CSV).to_excel(w, sheet_name="FDA_Drug_References", index=False)
        print(f"[WARN] Original locked; wrote {out_path}")

    # Light styling on FINAL_SUMARY
    wb = load_workbook(out_path)
    ws = wb["FINAL_SUMARY"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    lig_fill = PatternFill("solid", fgColor="C9A0DC")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        if cell.value in ("GtoP_top_ligands_per_gene", "recommended_drugs_per_gene"):
            cell.fill = lig_fill
            cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "C2"
    wb.save(out_path)
    print(f"[DONE] {out_path}")


if __name__ == "__main__":
    main()
